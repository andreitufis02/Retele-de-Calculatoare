from openpyxl import load_workbook
from flask import Flask, request, redirect

app = Flask(__name__)


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':

        # preluați datele din formular
        nume = request.form['nume']
        email = request.form['email']
        telefon = request.form['telefon']
        mesaj = request.form['mesaj']

        # deschideți fișierul Excel existent
        wb = load_workbook(filename='formular.xlsx')
        ws = wb.active

        # scrieți datele în fișierul Excel existent
        row = ws.max_row + 1
        ws.cell(row=row, column=1).value = nume
        ws.cell(row=row, column=2).value = email
        ws.cell(row=row, column=3).value = telefon
        ws.cell(row=row, column=4).value = mesaj
        wb.save('formular.xlsx')

        # redirecționați utilizatorul înapoi la formularul HTML
        return redirect('/')
    else:
        return '''
            <form method="post">
                Nume: <input type="text" name="nume"><br>
                Email: <input type="text" name="email"><br>
                Telefon: <input type="text" name="telefon"><br>
                Mesaj: <textarea name="mesaj"></textarea><br>
                <input type="submit" value="Trimite">
            </form>
        '''


if __name__ == '__main__':
    app.run()
